import asyncio
import pandas as pd
from typing import Dict, List, Optional, Any, Union
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
import io
import base64
import json
import logging
from pathlib import Path
from enum import Enum

# ReportLab para PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference, PieChart
from openpyxl.drawing.image import Image as ExcelImage

# Visualizaciones
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.graph_objects as go
import plotly.io as pio

# Cache y rate limiting
from fastapi_cache.decorator import cache
from slowapi import Limiter
from slowapi.util import get_remote_address

from ..core.config import settings
from ..services.analytics.analytics_engine import AnalyticsEngine
from ..services.email_automation import EmailAutomationService

# Logger
logger = logging.getLogger("report_generator")

class ReportFormat(str, Enum):
    PDF = "pdf"
    EXCEL = "excel"
    JSON = "json"
    HTML = "html"
    CSV = "csv"

class ReportType(str, Enum):
    EXECUTIVE = "executive"
    DETAILED_ANALYTICS = "detailed_analytics"
    CHANNEL_PERFORMANCE = "channel_performance"
    LEAD_QUALITY = "lead_quality"
    CUSTOM = "custom"

class ReportGenerator:
    """Generador de reportes en múltiples formatos (PDF, Excel, HTML, JSON)"""
    
    def __init__(self):
        self.analytics_engine = AnalyticsEngine()
        self.email_service = EmailAutomationService()
        
        # Configuración de estilos
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
        
        # Configuración de la compañía
        self.company_config = {
            'name': settings.COMPANY_NAME,
            'logo_path': settings.COMPANY_LOGO_PATH,
            'colors': {
                'primary': '#3B82F6',
                'secondary': '#10B981', 
                'accent': '#F59E0B',
                'warning': '#EF4444',
                'dark': '#1F2937',
                'light': '#F9FAFB'
            },
            'contact_info': settings.COMPANY_CONTACT_INFO
        }
        
        # Templates de reportes
        self.report_templates = {
            ReportType.EXECUTIVE: self._executive_template,
            ReportType.DETAILED_ANALYTICS: self._detailed_analytics_template,
            ReportType.CHANNEL_PERFORMANCE: self._channel_performance_template,
            ReportType.LEAD_QUALITY: self._lead_quality_template
        }
        
        # Rate limiting
        self.limiter = Limiter(key_func=get_remote_address)
        
        # Cache de reportes
        self.cache_enabled = settings.REPORT_CACHE_ENABLED
        
    def _setup_custom_styles(self):
        """Configura estilos personalizados para PDF"""
        self.custom_styles = {
            'title': ParagraphStyle(
                'CustomTitle',
                parent=self.styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor(self.company_config['colors']['primary']),
                spaceAfter=30,
                alignment=1  # Centered
            ),
            'subtitle': ParagraphStyle(
                'CustomSubtitle',
                parent=self.styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor(self.company_config['colors']['dark']),
                spaceAfter=20
            ),
            'kpi_value': ParagraphStyle(
                'KPIValue',
                parent=self.styles['Normal'],
                fontSize=18,
                textColor=colors.HexColor(self.company_config['colors']['primary']),
                alignment=1
            ),
            'kpi_label': ParagraphStyle(
                'KPILabel',
                parent=self.styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor(self.company_config['colors']['dark']),
                alignment=1
            )
        }
    
    @cache(expire=3600)  # Cache por 1 hora
    async def generate_report(self,
                            report_type: ReportType,
                            format_type: ReportFormat,
                            period: str = "monthly",
                            custom_filters: Optional[Dict] = None,
                            db: Session = None) -> Dict[str, Any]:
        """Genera reporte en el formato especificado"""
        
        try:
            start_time = datetime.utcnow()
            logger.info(f"Generando reporte {report_type} en formato {format_type}")
            
            # Validar parámetros
            self._validate_report_parameters(report_type, format_type, period)
            
            # Obtener datos del reporte
            report_data = await self._get_report_data(report_type, period, custom_filters, db)
            
            # Generar en formato específico
            if format_type == ReportFormat.PDF:
                content, filename, content_type = await self._generate_pdf(report_type, report_data, period)
            elif format_type == ReportFormat.EXCEL:
                content, filename, content_type = await self._generate_excel(report_type, report_data, period)
            elif format_type == ReportFormat.JSON:
                content, filename, content_type = await self._generate_json(report_type, report_data, period)
            elif format_type == ReportFormat.HTML:
                content, filename, content_type = await self._generate_html(report_type, report_data, period)
            elif format_type == ReportFormat.CSV:
                content, filename, content_type = await self._generate_csv(report_type, report_data, period)
            else:
                raise ValueError(f"Formato {format_type} no soportado")
            
            generation_time = (datetime.utcnow() - start_time).total_seconds()
            
            logger.info(f"Reporte generado exitosamente en {generation_time:.2f}s")
            
            return {
                "success": True,
                "filename": filename,
                "content": content,
                "content_type": content_type,
                "size_bytes": len(content),
                "generation_time_seconds": generation_time,
                "metadata": {
                    "report_type": report_type,
                    "format": format_type,
                    "period": period,
                    "generated_at": datetime.utcnow().isoformat(),
                    "cache_key": self._generate_cache_key(report_type, format_type, period, custom_filters)
                }
            }
            
        except Exception as e:
            logger.error(f"Error generando reporte: {str(e)}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "error_type": type(e).__name__
            }
    
    def _validate_report_parameters(self, report_type: ReportType, format_type: ReportFormat, period: str):
        """Valida los parámetros del reporte"""
        
        valid_periods = ["daily", "weekly", "monthly", "quarterly", "yearly"]
        if period not in valid_periods:
            raise ValueError(f"Período '{period}' no válido. Use: {', '.join(valid_periods)}")
        
        # Verificar formatos soportados por tipo de reporte
        supported_formats = self.get_report_metadata(report_type).get('supported_formats', [])
        if format_type.value not in supported_formats:
            raise ValueError(f"Formato {format_type} no soportado para reporte {report_type}")
    
    async def _get_report_data(self, report_type: ReportType, period: str, filters: Optional[Dict], db: Session) -> Dict:
        """Obtiene datos para el reporte especificado"""
        
        period_days = {
            'daily': 1,
            'weekly': 7,
            'monthly': 30,
            'quarterly': 90,
            'yearly': 365
        }.get(period, 30)
        
        template_func = self.report_templates.get(report_type)
        if not template_func:
            raise ValueError(f"Tipo de reporte '{report_type}' no soportado")
        
        return await template_func(period_days, filters, db)
    
    async def _generate_pdf(self, report_type: ReportType, report_data: Dict, period: str) -> tuple:
        """Genera reporte en formato PDF"""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=72,
            rightMargin=72,
            leftMargin=72,
            bottomMargin=72
        )
        
        story = self._build_pdf_story(report_type, report_data, period)
        doc.build(story)
        
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        filename = f"reporte_{report_type.value}_{period}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        return pdf_bytes, filename, "application/pdf"
    
    def _build_pdf_story(self, report_type: ReportType, report_data: Dict, period: str) -> List:
        """Construye el contenido del PDF"""
        
        story = []
        
        # Header con logo y información de la compañía
        story.extend(self._build_pdf_header())
        
        # Título del reporte
        title_text = f"Reporte {self.get_report_metadata(report_type)['name']}"
        story.append(Paragraph(title_text, self.custom_styles['title']))
        
        # Información del período
        period_text = f"Período: {period.title()} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        story.append(Paragraph(period_text, self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Contenido específico del reporte
        if report_type == ReportType.EXECUTIVE:
            story.extend(self._build_executive_pdf_content(report_data))
        elif report_type == ReportType.DETAILED_ANALYTICS:
            story.extend(self._build_detailed_pdf_content(report_data))
        elif report_type == ReportType.CHANNEL_PERFORMANCE:
            story.extend(self._build_channel_pdf_content(report_data))
        elif report_type == ReportType.LEAD_QUALITY:
            story.extend(self._build_quality_pdf_content(report_data))
        
        # Footer
        story.append(PageBreak())
        story.extend(self._build_pdf_footer())
        
        return story
    
    def _build_pdf_header(self) -> List:
        """Construye el header del PDF con logo e información de la compañía"""
        
        header_elements = []
        
        # Intentar cargar logo
        try:
            if Path(self.company_config['logo_path']).exists():
                logo = Image(self.company_config['logo_path'], width=2*inch, height=0.5*inch)
                logo.hAlign = 'LEFT'
                header_elements.append(logo)
        except:
            pass  # Continuar sin logo si hay error
        
        header_elements.extend([
            Spacer(1, 10),
            Paragraph(self.company_config['name'], self.styles['Heading2']),
            Paragraph("Reporte de Analytics", self.styles['Normal']),
            Spacer(1, 20)
        ])
        
        return header_elements
    
    def _build_pdf_footer(self) -> List:
        """Construye el footer del PDF"""
        
        return [
            Spacer(1, 20),
            Paragraph("Confidencial - Uso Interno", self.styles['Italic']),
            Paragraph(f"Generado por {self.company_config['name']}", self.styles['Italic']),
            Paragraph(f"Contacto: {self.company_config['contact_info']}", self.styles['Italic'])
        ]
    
    def _build_executive_pdf_content(self, report_data: Dict) -> List:
        """Construye contenido PDF para reporte ejecutivo"""
        
        content = []
        
        # Resumen ejecutivo
        content.append(Paragraph("Resumen Ejecutivo", self.custom_styles['subtitle']))
        content.append(Spacer(1, 12))
        
        # KPIs en formato tarjetas
        kpis = report_data.get('summary', {}).get('kpi_summary', {})
        if kpis:
            kpi_table_data = self._create_kpi_table_data(kpis)
            kpi_table = Table(kpi_table_data, colWidths=[150, 100, 80])
            kpi_table.setStyle(self._get_kpi_table_style())
            content.append(kpi_table)
            content.append(Spacer(1, 20))
        
        # Insights y recomendaciones
        content.extend(self._build_insights_section(report_data))
        
        return content
    
    def _create_kpi_table_data(self, kpis: Dict) -> List:
        """Crea datos para tabla de KPIs"""
        
        table_data = [['Métrica', 'Valor', 'Cambio']]
        
        kpi_formatters = {
            'conversion_rate': lambda v: f"{v:.1%}",
            'roi': lambda v: f"{v:.0f}%",
            'revenue_attributed': lambda v: f"${v:,.0f}",
            'cost_per_lead': lambda v: f"${v:.2f}",
            'response_time': lambda v: f"{v:.1f}s",
            'default': lambda v: f"{v:,.0f}"
        }
        
        for kpi_name, kpi_data in kpis.items():
            value = kpi_data.get('value', 0)
            change = kpi_data.get('change', 0)
            
            formatter = kpi_formatters.get(kpi_name, kpi_formatters['default'])
            formatted_value = formatter(value)
            formatted_change = f"{change:+.1f}%"
            
            table_data.append([
                kpi_name.replace('_', ' ').title(),
                formatted_value,
                formatted_change
            ])
        
        return table_data
    
    def _get_kpi_table_style(self) -> TableStyle:
        """Retorna estilo para tabla de KPIs"""
        
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.company_config['colors']['primary'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(self.company_config['colors']['light'])),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
        ])
    
    def _build_insights_section(self, report_data: Dict) -> List:
        """Construye sección de insights y recomendaciones"""
        
        content = []
        
        insights = report_data.get('insights', [])
        if insights:
            content.append(Paragraph("Insights Principales", self.custom_styles['subtitle']))
            content.append(Spacer(1, 12))
            
            for i, insight in enumerate(insights[:5], 1):
                insight_text = f"<b>{i}. {insight['title']}:</b> {insight['description']}"
                if insight.get('impact'):
                    insight_text += f" <i>(Impacto: {insight['impact']})</i>"
                
                content.append(Paragraph(insight_text, self.styles['Normal']))
                content.append(Spacer(1, 8))
        
        recommendations = report_data.get('recommendations', [])
        if recommendations:
            content.append(Spacer(1, 20))
            content.append(Paragraph("Recomendaciones", self.custom_styles['subtitle']))
            content.append(Spacer(1, 12))
            
            for i, rec in enumerate(recommendations[:5], 1):
                rec_text = f"<b>{i}. {rec['title']}:</b> {rec['description']}"
                if rec.get('priority'):
                    rec_text += f" <i>(Prioridad: {rec['priority']})</i>"
                
                content.append(Paragraph(rec_text, self.styles['Normal']))
                content.append(Spacer(1, 8))
        
        return content
    
    async def _generate_excel(self, report_type: ReportType, report_data: Dict, period: str) -> tuple:
        """Genera reporte en formato Excel"""
        
        workbook = Workbook()
        
        # Hoja de resumen
        ws_summary = workbook.active
        ws_summary.title = "Resumen"
        self._build_excel_summary_sheet(ws_summary, report_type, report_data, period)
        
        # Hojas adicionales según tipo de reporte
        if report_type == ReportType.EXECUTIVE:
            self._build_excel_kpi_sheet(workbook, report_data)
            self._build_excel_insights_sheet(workbook, report_data)
        elif report_type == ReportType.CHANNEL_PERFORMANCE:
            self._build_excel_channels_sheet(workbook, report_data)
            self._build_excel_attribution_sheet(workbook, report_data)
        
        # Guardar en buffer
        buffer = io.BytesIO()
        workbook.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        filename = f"reporte_{report_type.value}_{period}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return excel_bytes, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    def _build_excel_summary_sheet(self, worksheet, report_type: ReportType, report_data: Dict, period: str):
        """Construye hoja de resumen en Excel"""
        
        # Estilos
        title_font = Font(size=16, bold=True, color="FFFFFF")
        header_font = Font(bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color=self.company_config['colors']['primary'], fill_type="solid")
        header_fill = PatternFill(start_color=self.company_config['colors']['secondary'], fill_type="solid")
        
        # Título
        worksheet.merge_cells('A1:D1')
        worksheet['A1'] = f"Reporte {self.get_report_metadata(report_type)['name']}"
        worksheet['A1'].font = title_font
        worksheet['A1'].fill = title_fill
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        worksheet['A2'] = f"Período: {period.title()}"
        worksheet['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # KPIs principales
        kpis = report_data.get('summary', {}).get('kpi_summary', {})
        if kpis:
            # Headers
            headers = ['Métrica', 'Valor Actual', 'Valor Anterior', 'Cambio %']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=5, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            row = 6
            for kpi_name, kpi_data in kpis.items():
                worksheet.cell(row=row, column=1).value = kpi_name.replace('_', ' ').title()
                worksheet.cell(row=row, column=2).value = kpi_data.get('value', 0)
                worksheet.cell(row=row, column=3).value = kpi_data.get('previous_value', 0)
                worksheet.cell(row=row, column=4).value = kpi_data.get('change', 0)
                row += 1
            
            # Ajustar anchos de columna
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 12
    
    async def _generate_json(self, report_type: ReportType, report_data: Dict, period: str) -> tuple:
        """Genera reporte en formato JSON"""
        
        json_data = {
            "metadata": {
                "report_type": report_type.value,
                "period": period,
                "generated_at": datetime.utcnow().isoformat(),
                "version": "1.0"
            },
            "data": report_data
        }
        
        json_bytes = json.dumps(json_data, indent=2, ensure_ascii=False).encode('utf-8')
        filename = f"reporte_{report_type.value}_{period}_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        
        return json_bytes, filename, "application/json"
    
    async def _generate_html(self, report_type: ReportType, report_data: Dict, period: str) -> tuple:
        """Genera reporte en formato HTML"""
        
        html_template = self._get_html_template(report_type)
        html_content = html_template.format(
            report_title=self.get_report_metadata(report_type)['name'],
            period=period,
            generated_date=datetime.now().strftime('%d/%m/%Y %H:%M'),
            company_name=self.company_config['name'],
            report_data=json.dumps(report_data, indent=2)
        )
        
        html_bytes = html_content.encode('utf-8')
        filename = f"reporte_{report_type.value}_{period}_{datetime.now().strftime('%Y%m%d_%H%M')}.html"
        
        return html_bytes, filename, "text/html"
    
    async def _generate_csv(self, report_type: ReportType, report_data: Dict, period: str) -> tuple:
        """Genera reporte en formato CSV"""
        
        # Convertir datos a DataFrame y luego a CSV
        flat_data = self._flatten_report_data(report_data)
        df = pd.DataFrame([flat_data])
        
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        csv_bytes = csv_buffer.getvalue().encode('utf-8')
        csv_buffer.close()
        
        filename = f"reporte_{report_type.value}_{period}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        
        return csv_bytes, filename, "text/csv"
    
    def _flatten_report_data(self, data: Dict, parent_key: str = '', sep: str = '_') -> Dict:
        """Aplana datos anidados para CSV"""
        items = []
        for k, v in data.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_report_data(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Para listas, convertimos a string JSON
                items.append((new_key, json.dumps(v)))
            else:
                items.append((new_key, v))
        return dict(items)
    
    def _get_html_template(self, report_type: ReportType) -> str:
        """Retorna template HTML para el reporte"""
        
        return f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <title>Reporte {self.get_report_metadata(report_type)['name']}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                .header {{ background-color: {self.company_config['colors']['primary']}; color: white; padding: 20px; }}
                .content {{ margin: 20px 0; }}
                .kpi {{ border: 1px solid #ddd; padding: 15px; margin: 10px 0; }}
                .positive {{ color: green; }}
                .negative {{ color: red; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{{report_title}}</h1>
                <p>Período: {{period}} | Generado: {{generated_date}}</p>
                <p>Compañía: {{company_name}}</p>
            </div>
            <div class="content">
                <pre>{{report_data}}</pre>
            </div>
        </body>
        </html>
        """
    
    async def generate_and_send_report(self,
                                     report_type: ReportType,
                                     format_type: ReportFormat,
                                     period: str,
                                     email_recipients: List[str],
                                     custom_filters: Optional[Dict] = None,
                                     db: Session = None) -> Dict[str, Any]:
        """Genera y envía reporte por email"""
        
        try:
            # Generar reporte
            result = await self.generate_report(report_type, format_type, period, custom_filters, db)
            
            if not result['success']:
                return result
            
            # Enviar por email
            email_result = await self.email_service.send_report_email(
                recipients=email_recipients,
                subject=f"Reporte {self.get_report_metadata(report_type)['name']} - {period.title()}",
                body=self._get_email_body(report_type, period),
                attachment_content=result['content'],
                attachment_filename=result['filename'],
                attachment_content_type=result['content_type'],
                db=db
            )
            
            result['email_result'] = email_result
            return result
            
        except Exception as e:
            logger.error(f"Error enviando reporte por email: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def _get_email_body(self, report_type: ReportType, period: str) -> str:
        """Genera cuerpo del email para el reporte"""
        
        metadata = self.get_report_metadata(report_type)
        
        return f"""
        <h2>Reporte {metadata['name']}</h2>
        <p>Se adjunta el reporte {metadata['name']} para el período {period}.</p>
        <p><strong>Descripción:</strong> {metadata['description']}</p>
        <p>Este reporte fue generado automáticamente el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}.</p>
        <p>Si tiene alguna pregunta, no dude en contactarnos.</p>
        <br>
        <p>Saludos cordiales,<br>{self.company_config['name']}</p>
        """
    
    def get_report_metadata(self, report_type: ReportType) -> Dict[str, Any]:
        """Obtiene metadatos de un tipo de reporte"""
        
        metadata = {
            ReportType.EXECUTIVE: {
                'name': 'Ejecutivo',
                'description': 'Resumen ejecutivo con KPIs principales y insights',
                'sections': ['kpis', 'insights', 'recommendations'],
                'estimated_generation_time': '30-60 segundos',
                'supported_formats': ['json', 'pdf', 'excel', 'html'],
                'audience': 'Ejecutivos, Gerentes'
            },
            ReportType.DETAILED_ANALYTICS: {
                'name': 'Analytics Detallado',
                'description': 'Análisis profundo de todas las métricas y tendencias',
                'sections': ['funnel', 'quality', 'engagement', 'trends'],
                'estimated_generation_time': '60-90 segundos',
                'supported_formats': ['json', 'pdf', 'excel'],
                'audience': 'Analistas, Equipo de Marketing'
            },
            ReportType.CHANNEL_PERFORMANCE: {
                'name': 'Performance por Canal',
                'description': 'Análisis detallado de ROI y performance por canal',
                'sections': ['channels', 'attribution', 'costs', 'trends'],
                'estimated_generation_time': '45-75 segundos',
                'supported_formats': ['json', 'pdf', 'excel', 'csv'],
                'audience': 'Equipo de Marketing, Media Buyers'
            },
            ReportType.LEAD_QUALITY: {
                'name': 'Calidad de Leads',
                'description': 'Análisis de scoring, segmentación y calidad',
                'sections': ['scoring', 'sources', 'trends', 'recommendations'],
                'estimated_generation_time': '30-45 segundos',
                'supported_formats': ['json', 'pdf', 'excel'],
                'audience': 'Equipo de Ventas, Marketing'
            }
        }
        
        return metadata.get(report_type, {})
    
    def get_available_reports(self) -> List[Dict[str, Any]]:
        """Retorna lista de reportes disponibles"""
        
        reports = []
        for report_type in ReportType:
            metadata = self.get_report_metadata(report_type)
            reports.append({
                'type': report_type.value,
                'name': metadata['name'],
                'description': metadata['description'],
                'supported_formats': metadata['supported_formats'],
                'audience': metadata['audience'],
                'estimated_generation_time': metadata['estimated_generation_time']
            })
        
        return reports
    
    def _generate_cache_key(self, report_type: ReportType, format_type: ReportFormat, 
                          period: str, filters: Optional[Dict]) -> str:
        """Genera clave única para cache"""
        
        filter_str = json.dumps(filters, sort_keys=True) if filters else "no_filters"
        return f"report:{report_type.value}:{format_type.value}:{period}:{filter_str}"
    
    async def schedule_recurring_report(self, 
                                      report_config: Dict,
                                      schedule: str,
                                      recipients: List[str],
                                      db: Session = None) -> Dict[str, Any]:
        """Programa reporte recurrente"""
        
        try:
            # Validar configuración
            required_fields = ['type', 'format', 'period']
            for field in required_fields:
                if field not in report_config:
                    raise ValueError(f"Campo requerido faltante: {field}")
            
            report_type = ReportType(report_config['type'])
            format_type = ReportFormat(report_config['format'])
            
            # Crear entrada de reporte programado
            scheduled_report = {
                "id": f"sched_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
                "report_type": report_type.value,
                "format": format_type.value,
                "period": report_config['period'],
                "schedule": schedule,
                "recipients": recipients,
                "filters": report_config.get('filters', {}),
                "next_execution": self._calculate_next_execution(schedule),
                "is_active": True,
                "created_at": datetime.utcnow().isoformat(),
                "created_by": "api"
            }
            
            # En producción, guardar en base de datos
            # await self._save_scheduled_report(scheduled_report, db)
            
            logger.info(f"Reporte recurrente programado: {scheduled_report['id']}")
            
            return {
                "success": True,
                "scheduled_report": scheduled_report,
                "message": f"Reporte programado {schedule} creado exitosamente"
            }
            
        except Exception as e:
            logger.error(f"Error programando reporte recurrente: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def _calculate_next_execution(self, schedule: str) -> str:
        """Calcula próxima ejecución según schedule"""
        
        now = datetime.utcnow()
        
        if schedule == 'daily':
            next_exec = now.replace(hour=8, minute=0, second=0, microsecond=0)
            if next_exec <= now:
                next_exec += timedelta(days=1)
        elif schedule == 'weekly':
            days_until_monday = (7 - now.weekday()) % 7
            next_exec = now.replace(hour=8, minute=0, second=0, microsecond=0) + timedelta(days=days_until_monday)
        elif schedule == 'monthly':
            next_exec = now.replace(day=1, hour=8, minute=0, second=0, microsecond=0)
            if next_exec <= now:
                next_exec = (next_exec + timedelta(days=32)).replace(day=1)
        else:
            raise ValueError(f"Schedule no soportado: {schedule}")
        
        return next_exec.isoformat()
    
    # Métodos de template (simplificados para el ejemplo)
    async def _executive_template(self, days: int, filters: Optional[Dict], db: Session) -> Dict:
        return await self.analytics_engine.generate_executive_report(
            "monthly" if days >= 30 else "weekly", db
        )
    
    async def _detailed_analytics_template(self, days: int, filters: Optional[Dict], db: Session) -> Dict:
        return {"data": "detailed_analytics_placeholder"}
    
    async def _channel_performance_template(self, days: int, filters: Optional[Dict], db: Session) -> Dict:
        return {"data": "channel_performance_placeholder"}
    
    async def _lead_quality_template(self, days: int, filters: Optional[Dict], db: Session) -> Dict:
        return {"data": "lead_quality_placeholder"}
    
    # Métodos de construcción de contenido PDF (simplificados)
    def _build_detailed_pdf_content(self, report_data: Dict) -> List:
        return [Paragraph("Contenido detallado placeholder", self.styles['Normal'])]
    
    def _build_channel_pdf_content(self, report_data: Dict) -> List:
        return [Paragraph("Contenido canales placeholder", self.styles['Normal'])]
    
    def _build_quality_pdf_content(self, report_data: Dict) -> List:
        return [Paragraph("Contenido calidad placeholder", self.styles['Normal'])]
    
    # Métodos de construcción de Excel (simplificados)
    def _build_excel_kpi_sheet(self, workbook, report_data: Dict):
        pass
    
    def _build_excel_insights_sheet(self, workbook, report_data: Dict):
        pass
    
    def _build_excel_channels_sheet(self, workbook, report_data: Dict):
        pass
    
    def _build_excel_attribution_sheet(self, workbook, report_data: Dict):
        pass

# FastAPI Router para reportes
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, Query
from ..core.database import get_db

router = APIRouter()
report_generator = ReportGenerator()

@router.get("/reports/available")
async def get_available_reports():
    """Obtiene lista de reportes disponibles"""
    return report_generator.get_available_reports()

@router.post("/reports/generate")
async def generate_report(
    report_type: ReportType,
    format_type: ReportFormat,
    period: str = Query("monthly", regex="^(daily|weekly|monthly|quarterly|yearly)$"),
    custom_filters: Optional[Dict] = None,
    db: Session = Depends(get_db)
):
    """Genera un reporte específico"""
    return await report_generator.generate_report(report_type, format_type, period, custom_filters, db)

@router.post("/reports/send")
async def send_report(
    report_type: ReportType,
    format_type: ReportFormat,
    period: str,
    email_recipients: List[str],
    custom_filters: Optional[Dict] = None,
    db: Session = Depends(get_db)
):
    """Genera y envía reporte por email"""
    return await report_generator.generate_and_send_report(
        report_type, format_type, period, email_recipients, custom_filters, db
    )

@router.post("/reports/schedule")
async def schedule_report(
    report_config: Dict,
    schedule: str = Query(..., regex="^(daily|weekly|monthly)$"),
    recipients: List[str] = [],
    db: Session = Depends(get_db)
):
    """Programa un reporte recurrente"""
    return await report_generator.schedule_recurring_report(report_config, schedule, recipients, db)